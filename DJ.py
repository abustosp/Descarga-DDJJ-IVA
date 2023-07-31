import tabula
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font

# Leer el Excel que posee las ubicaciones
Excel = pd.read_excel('Listado DDJJ IVA.xlsx')

# Reemplazar los \ por / en la columna 'Ubicación Descarga'
Excel['Ubicación Descarga'] = Excel['Ubicación Descarga'].astype(str).str.replace("\\", "/")

# Rellenar los np.nan de 'importar' con 'No'
Excel['Importar'] = Excel['Importar'].fillna('No')

# Filtrar los datos que tengan 'Si' o 'SI' en la columna 'Importar'
Excel = Excel[Excel['Importar'].str.contains('Si|SI')]

# hacer un for con los items del Excel
for i in range(len(Excel)):

    # Leer el PDF concatenando las columnas 'Ubicación Descarga' y 'DDJJ IVA'
    df = tabula.read_pdf((Excel['Ubicación Descarga'][i] + Excel['DDJJ IVA'][i] + ".pdf"), pages='all')

    Saldos = df[1]

    #Convertir la primera fila en el encabezado
    Saldos.columns = Saldos.iloc[0]

    #Eliminar la primera fila
    Saldos.drop(Saldos.index[0], inplace=True)

    Saldos['Importe'] = Saldos['Importe'].str.replace('$ ', '').astype(float)

    Ubicacon_Excel = Excel['Ubicación Descarga'][i] + Excel['DDJJ IVA'][i] + ".xlsx"

    Saldos.to_excel(Ubicacon_Excel, index=False , sheet_name='DDJJ IVA')


    # Aplicar Formato a los Excels

    workbook = openpyxl.load_workbook(Ubicacon_Excel)
    hoja1 = workbook['DDJJ IVA']  # Nombre de la hoja del DataFrame

    # Darle formato a los Títulos de las columnas
    Fondotitulo = PatternFill(start_color='002060' , end_color='002060' ,  fill_type='solid')
    LetraColor = Font(color='FFFFFF')

    # Aplicar formato al encabezado
    for cell in hoja1[1]:
        cell.fill = Fondotitulo
        cell.font = LetraColor

    # Autoajustar los anchos de las columnas según el contenido
    for column_cells in hoja1.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        hoja1.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Agregar filtros de datos de ambas hojas
    hoja1.auto_filter.ref = hoja1.dimensions

    # Guardar el archivo Excel
    workbook.save(Ubicacon_Excel)